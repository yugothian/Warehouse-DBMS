import tkinter as tk
from tkinter import ttk
from tkinter import *
from tkcalendar import DateEntry
from PIL import ImageTk, Image
from tkinter import filedialog
import openpyxl
from openpyxl import Workbook
from tkinter import messagebox
import pathlib
import cv2
import numpy
import pandas as pd

class Visualize(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Management System")
        self.geometry("350x400")

        self.label = ttk.Label(self, text="Warehouse Management System", font=24)
        self.label.place(x=60, y=30)

        self.add_raw = ttk.Button(self, text="Add Raw Materials", command = addRaw)
        self.add_raw.place(x=50, y=100, width=120, height=40)

        self.edit_raw = ttk.Button(self, text="Edit Raw Materials", command=editRaw)
        self.edit_raw.place(x=50, y=150, width=120, height=40)

        self.remove_raw = ttk.Button(self, text="Remove Raw Materials", command=removeRaw)
        self.remove_raw.place(x=50, y=200, width=120, height=40)

        self.chekccam = ttk.Button(self, text="Control Cams", command=controlCams)
        self.chekccam.place(x=50, y=250, width=120, height=40)

        self.add_product = ttk.Button(self, text="Add Product", command=addProduct)
        self.add_product.place(x=200, y=100, width=120, height=40)

        self.edit_product = ttk.Button(self, text="Edit Product",command=editProduct)
        self.edit_product.place(x=200, y=150, width=120, height=40)

        self.remove_product = ttk.Button(self, text="Remove Product", command=removeProduct)
        self.remove_product.place(x=200, y=200, width=120, height=40)

        self.list_product = ttk.Button(self, text="List Items", command=listRaw_Product) 
        self.list_product.place(x=200, y=250, width=120, height=40)


class addRaw(tk.Toplevel):
    def __init__(self):
        super().__init__()

        self.title("Add Raw")
        self.geometry("600x800")

        self.name_label = ttk.Label(self, text="Name:", font=("Ariel",12))
        self.name_label.place(x=130, y=50)

        self.name_entry = ttk.Entry(self)
        self.name_entry.place (x=200, y=50)

        self.storage_code_label = ttk.Label(self, text="Storage Code:", font=("Ariel",12))
        self.storage_code_label.place(x=72, y=75)

        self.storage_code_entry = ttk.Entry(self)
        self.storage_code_entry.place (x=200, y=75)   

        self.purchase_date_label = ttk.Label(self, text="Date of Purchase:", font=("Ariel",12))
        self.purchase_date_label.place(x=50, y=100)

        self.purchase_date_entry = DateEntry(self, selectmode='day')
        self.purchase_date_entry.place(x=200, y=100)

        self.supplier_name_label = ttk.Label(self, text="Name of Supplier:", font=("Ariel",12))
        self.supplier_name_label.place(x=50, y= 125)

        self.supplier_name_entry = ttk.Entry(self)
        self.supplier_name_entry.place(x=200, y=125)
        
        self.expiration_date_label = ttk.Label(self, text="Date of Expiration: ", font=("Ariel",12))
        self.expiration_date_label.place(x=45, y=150)

        self.expiration_date_entry = DateEntry(self, selectmode='day')
        self.expiration_date_entry.place(x=200, y=150)

        self.description_label = ttk.Label(self, text="Description:", font=("Ariel",12))
        self.description_label.place(x=90, y=175)

        self.description_entry = Text(self, height = 7, width = 40)
        self.description_entry.place(x=200, y=175)

        self.add_image_button = ttk.Button(self, text="Add Image", command= self.open_image)
        self.add_image_button.place(x=100, y=300)

        self.add_raw = ttk.Button(self,text="Add", command= self.add_to_excel)
        self.add_raw.place(x=240, y=650,width=120, height=40)


    def open_image(self):     

        path = filedialog.askopenfilename(filetypes=[("Image File",'.jpg')])
        im = Image.open(path)
        self.tkimage = ImageTk.PhotoImage(im)
        myvar=Label(self, image = self.tkimage, width=300, height=300, background="gray")
        myvar.image = self.tkimage
        myvar.place(x=200, y=300)

    def add_to_excel(self):
        self.raw_materials = pathlib.Path("raw_materials.xlsx")

        if self.raw_materials.exists():
            pass
        else:
            self.raw_materials = Workbook()
            self.sheet = self.raw_materials.active
            self.sheet["A1"] = "Storage Code"
            self.sheet["B1"] = "Name"
            self.sheet["C1"] = "Date of Purchase"
            self.sheet["D1"] = "Name of Supplier"
            self.sheet["E1"] = "Storage Expiration Date"
            self.sheet["F1"] = "Description"
            self.sheet["G1"] = "Image"

            self.raw_materials.save("raw_materials.xlsx")

        
        self.name = self.name_entry.get()
        self.storage =self.storage_code_entry.get()
        self.supplier = self.supplier_name_entry.get()
        self.purchase_date = str(self.purchase_date_entry.get_date())
        self.expiration = str(self.expiration_date_entry.get_date())
        self.description = self.description_entry.get("1.0",'end-1c')


        if (self.name == '') or (self.storage == '') or (self.supplier == '') or (self.description == ''):

           self.warning = messagebox.showwarning("Missing Field", "One or More Fields are Empty, Please Check Again")
    
        else:
            
            self.raw_materials = openpyxl.load_workbook("raw_materials.xlsx")
            self.sheet = self.raw_materials.active

            self.sheet.cell(column=1, row=self.sheet.max_row+1, value=self.storage)
            self.sheet.cell(column=2, row=self.sheet.max_row, value=self.name)
            self.sheet.cell(column=3, row=self.sheet.max_row, value=self.purchase_date)
            self.sheet.cell(column=4, row=self.sheet.max_row, value=self.supplier)
            self.sheet.cell(column=5, row=self.sheet.max_row, value=self.expiration)
            self.sheet.cell(column=6, row=self.sheet.max_row, value=self.description)
            # self.sheet.cell(column=7, row=self.sheet.max_row, value=self.tkimage)
            # self.raw_image = Image(self.tkimage)
            # self.sheet.add_image(self.raw_image, "G1")


            self.raw_materials.save("raw_materials.xlsx")


class editRaw(tk.Toplevel):
    def __init__(self):
        super().__init__()

        self.geometry("1500x200")
        df = pd.read_excel("raw_materials.xlsx")

        n_rows = df.shape[0]
        n_cols = df.shape[1]

        column_names = df.columns
        i=0
        for j, col in enumerate(column_names):
            text = Text(self, width=16, height=1, bg = "#9BC2E6")
            text.grid(row=i,column=j)
            text.insert(INSERT, col)
            

        cells = {}

        for i in range(n_rows):
            for j in range(n_cols):
                text = Text(self, width=16, height=1)
                text.grid(row=i+1,column=j)
                text.insert(INSERT, df.loc[i][j])
                cells[(i,j)] = text		
                
        def do_something():

            for i in range(n_rows):
                for j in range(n_cols-1):
                    if df.loc[i][j] != cells[(i,j)].get("1.0", "end-1c"):
                        df.loc[[i],column_names[j]] = cells[(i,j)].get("1.0", "end-1c")
            df.to_excel("raw_materials.xlsx")


        save_button = Button(
            self, height = 2,
            width = 16,
            text ="Save",
            command = lambda:do_something())
        save_button.grid(row=7,column = 0)


        


    def edit_raw(self):

        self.book = openpyxl.load_workbook("raw_materials.xlsx", read_only=True)
        self.sheet = self.book.active

        for row in self.sheet.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):
            for cell in row:
                if cell.value == self.storage_code_entry.get():
                    
                    print(cell.value)
                    self.name_entry.insert(self.sheet.cell)
                    self.purchase_date_entry.config(self.purchase_date_entry.get_date())
                    
                    # self.new_data = [

                    #     self.storage_code_entry.get(),
                    #     self.name_entry.get(),
                    #     str(self.purchase_date_entry.get_date()),
                    #     self.supplier_name_entry.get(),
                    #     str(self.expiration_date_entry.get_date()),
                    #     self.description_entry.get()

                    # ]

                    self.sheet.cell(column=1, row=self.sheet.max_row+1, value=self.new_data[0])
                    self.sheet.cell(column=2, row=self.sheet.max_row, value=self.new_data[1])
                    self.sheet.cell(column=3, row=self.sheet.max_row, value=self.new_data[2])
                    self.sheet.cell(column=4, row=self.sheet.max_row, value=self.new_data[3])
                    self.sheet.cell(column=5, row=self.sheet.max_row, value=self.new_data[4])

                    self.raw_materials.save("raw_materials.xlsx")


class removeRaw(tk.Toplevel):
    def __init__(self):
        super().__init__()

        self.title("Remove Raw Materials")
        self.geometry("400x200")

        self.label = ttk.Label(self, text="Storage Code of the Material\n You Want To Remove", font=("Ariel", 12))
        self.label.place(x=20,y=20)

        self.remove_entry = ttk.Entry(self)
        self.remove_entry.place(x=240, y=30)

        self.rm_button = ttk.Button(self, text="Remove", command=self.remove)
        self.rm_button.place(x=250, y=70, width=100, height=30)

    def remove(self):
        self.book = openpyxl.load_workbook("raw_materials.xlsx")
        self.sheet = self.book.active

        self.s_code = self.remove_entry.get()
        self.row_count = self.sheet.max_row


        for row in self.sheet.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):
            for cell in row:
                if cell.value == self.s_code:
                    self.sheet.delete_rows(idx=cell.row)
                    self.book.save("raw_materials.xlsx")

                # else:
                #     self.warning = messagebox.showwarning("Invalid Code", "There's no item with that code, try again")


class listRaw_Product(tk.Toplevel):
    def __init__(self):
        super().__init__()
        
        self.title("List Raw Materials")
        self.geometry("1500x350")

        self.style = ttk.Style()
        self.style.theme_use("clam")

        self.frame = ttk.Frame(self)
        self.frame.pack(pady=20)
        
        self.label = Label(self, text='')
        self.label.pack(pady=20)
        self.tree = ttk.Treeview(self.frame)

        m = Menu(self)
        self.config(menu=m)

        file_menu = Menu(m, tearoff=False)
        m.add_cascade(label="List", menu=file_menu)
        file_menu.add_command(label="List Products", command=self.list_products)
        file_menu.add_command(label="List Raw Materials", command=self.list_raws)

    def list_raws(self):

        df = pd.read_excel("raw_materials.xlsx")

        self.clear_treeview()

        self.tree["column"] = list(df.columns)
        self.tree["show"] = "headings"

        for col in self.tree["column"]:
            self.tree.heading(col, text=col)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tree.insert("", "end", values=row)

        self.tree.pack()

    def list_products(self):

        df = pd.read_excel("products.xlsx")

        self.clear_treeview()

        self.tree["column"] = list(df.columns)
        self.tree["show"] = "headings"

        for col in self.tree["column"]:
            self.tree.heading(col, text=col)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tree.insert("", "end", values=row)

        self.tree.pack()

    def clear_treeview(self):
        self.tree.delete(*self.tree.get_children())


    
class controlCams(tk.Toplevel):
    def __init__(self):
        super().__init__()
        self.cap = cv2.VideoCapture(0)
        width, height = 600, 400
        self.cap.set(cv2.CAP_PROP_FRAME_WIDTH, width)
        self.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, height)

        self.title("Control Cameras")
        self.geometry("800x500")
        self.camera_label = ttk.Label(self)
        self.camera_label.pack(side=LEFT)

        button1 = ttk.Button(self, text="Open Camera_Warehouse", command=self.show_frames)
        button1.place(x=660, y=70, width=125, height=40)

        
        button1 = ttk.Button(self, text="Open Camera_Garage")
        button1.place(x=660, y=120, width=125, height=40)


        button1 = ttk.Button(self, text="Open Camera_Offices")
        button1.place(x=660, y=170, width=125, height=40)


    def show_frames(self):
        
        cv2image= cv2.cvtColor(self.cap.read()[1],cv2.COLOR_BGR2RGB)
        img = Image.fromarray(cv2image)
        
        imgtk = ImageTk.PhotoImage(image = img)
        self.camera_label.imgtk = imgtk
        self.camera_label.configure(image=imgtk)
        
        self.camera_label.after(20, self.show_frames)


class addProduct(tk.Toplevel):
    def __init__(self):
        super().__init__()
        
        self.book = openpyxl.load_workbook("raw_materials.xlsx", read_only=True)
        self.sheet = self.book.active
        
        self.title("Add Product")
        self.geometry("600x800")

        self.name_label = ttk.Label(self, text="Name:", font=("Ariel",12))
        self.name_label.place(x=200, y=50)

        self.name_entry = ttk.Entry(self)
        self.name_entry.place (x=270, y=50)

        self.storage_code_label = ttk.Label(self, text="Storage Code:", font=("Ariel",12))
        self.storage_code_label.place(x=142, y=75)

        self.storage_code_entry = ttk.Entry(self)
        self.storage_code_entry.place (x=270, y=75)   

        self.purchase_date_label = ttk.Label(self, text="Date of Purchase:", font=("Ariel",12))
        self.purchase_date_label.place(x=120, y=100)

        self.purchase_date_entry = DateEntry(self, selectmode='day')
        self.purchase_date_entry.place(x=270, y=100)

        self.customer_name_label = ttk.Label(self, text="Name of Customer:", font=("Ariel",12))
        self.customer_name_label.place(x=110, y= 125)

        self.customer_name_entry = ttk.Entry(self)
        self.customer_name_entry.place(x=270, y=125)
        
        self.expiration_date_label = ttk.Label(self, text="Date of Expiration: ", font=("Ariel",12))
        self.expiration_date_label.place(x=115, y=150)
        
        self.expiration_date_entry = DateEntry(self, selectmode='day')
        self.expiration_date_entry.place(x=270, y=150)

        self.used_raws_label = ttk.Label(self, text="List of Raws That Are Used: ", font=("Ariel",12))
        self.used_raws_label.place(x=50, y=175)


        # for i in self.sheet.column
        # self.used_raws_entry = ttk.Entry(self)
        # self.used_raws_entry.place(x=270, y=175)

        self.description_label = ttk.Label(self, text="Description:", font=("Ariel",12))
        self.description_label.place(x=160, y=200)

        self.description_entry = Text(self, height = 7, width = 30)
        self.description_entry.place(x=270, y=200)

        self.add_image_button = ttk.Button(self, text="Add Image", command= self.open_image)
        self.add_image_button.place(x=100, y=350)

        self.add_Product = ttk.Button(self,text="Add", command= self.add_to_excel)
        self.add_Product.place(x=240, y=700,width=120, height=40)


    def open_image(self):     

        path = filedialog.askopenfilename(filetypes=[("Image File",'.jpg')])
        im = Image.open(path)
        self.tkimage = ImageTk.PhotoImage(im)
        myvar=Label(self, image = self.tkimage, width=300, height=300, background="gray")
        myvar.image = self.tkimage
        myvar.place(x=200, y=350)

    def add_to_excel(self):
        self.products = pathlib.Path("products.xlsx")

        if self.products.exists():
            pass
        else:
            self.products = Workbook()
            self.sheet = self.products.active
            self.sheet["A1"] = "Storage Code"
            self.sheet["B1"] = "Name"
            self.sheet["C1"] = "Date of Purchase"
            self.sheet["D1"] = "Name of Customer"
            self.sheet["E1"] = "Storage Expiration Date"
            self.sheet["F1"] = "Description"
            self.sheet["G1"] = "Image"

            self.products.save("products.xlsx")

        
        self.name = self.name_entry.get()
        self.storage =self.storage_code_entry.get()
        self.customer = self.customer_name_entry.get()
        self.purchase_date = str(self.purchase_date_entry.get_date())
        self.expiration = str(self.expiration_date_entry.get_date())
        self.description = self.description_entry.get("1.0",'end-1c')


        if (self.name == '') or (self.storage == '') or (self.customer == '') or (self.description == ''):

           self.warning = messagebox.showwarning("Missing Field", "One or More Fields are Empty, Please Check Again")
    
        else:
            
            self.products = openpyxl.load_workbook("products.xlsx")
            self.sheet = self.products.active

            self.sheet.cell(column=1, row=self.sheet.max_row+1, value=self.storage)
            self.sheet.cell(column=2, row=self.sheet.max_row, value=self.name)
            self.sheet.cell(column=3, row=self.sheet.max_row, value=self.purchase_date)
            self.sheet.cell(column=4, row=self.sheet.max_row, value=self.customer)
            self.sheet.cell(column=5, row=self.sheet.max_row, value=self.expiration)
            self.sheet.cell(column=6, row=self.sheet.max_row, value=self.description)
            # self.sheet.cell(column=7, row=self.sheet.max_row, value=self.tkimage)
            # self.img = openpyxl.dProducting.image.self.tkimage
            # self.sheet
            

            self.products.save("products.xlsx")

class editProduct(tk.Toplevel):
    def __init__(self):
        super().__init__()

        self.geometry("1500x200")
        df = pd.read_excel("products.xlsx")

        n_rows = df.shape[0]
        n_cols = df.shape[1]

        column_names = df.columns
        i=0
        for j, col in enumerate(column_names):
            text = Text(self, width=16, height=1, bg = "#9BC2E6")
            text.grid(row=i,column=j)
            text.insert(INSERT, col)
            

        cells = {}

        for i in range(n_rows):
            for j in range(n_cols):
                text = Text(self, width=16, height=1)
                text.grid(row=i+1,column=j)
                text.insert(INSERT, df.loc[i][j])
                cells[(i,j)] = text		
                
        def do_something():

            for i in range(n_rows):
                for j in range(n_cols-1):
                    if df.loc[i][j] != cells[(i,j)].get("1.0", "end-1c"):
                        df.loc[[i],column_names[j]] = cells[(i,j)].get("1.0", "end-1c")
            df.to_excel("products.xlsx")


        save_button = Button(
            self, height = 2,
            width = 16,
            text ="Save",
            command = lambda:do_something())
        save_button.grid(row=7,column = 0)



class removeProduct(tk.Toplevel):
    def __init__(self):
        super().__init__()

        self.title("Remove Products")
        self.geometry("400x200")

        self.label = ttk.Label(self, text="Storage Code of the Product\n You Want To Remove", font=("Ariel", 12))
        self.label.place(x=20,y=20)

        self.remove_entry = ttk.Entry(self)
        self.remove_entry.place(x=240, y=30)

        self.rm_button = ttk.Button(self, text="Remove", command=self.remove)
        self.rm_button.place(x=250, y=70, width=100, height=30)

    def remove(self):
        self.book = openpyxl.load_workbook("products.xlsx")
        self.sheet = self.book.active

        self.s_code = self.remove_entry.get()
        self.row_count = self.sheet.max_row


        for row in self.sheet.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):
            for cell in row:
                if cell.value == self.s_code:
                    self.sheet.delete_rows(idx=cell.row)
                    self.book.save("products.xlsx")


if __name__ == "__main__":
    root = Visualize()
    root.mainloop()


